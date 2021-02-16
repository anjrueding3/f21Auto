from openpyxl import Workbook, load_workbook
import os, PyPDF2, re
import readPDF_forever21 as read


#parsedData = extractPoIHD_Dictionary('.')

directory = #commandLineArgs[1]

parsedData = read.extractPoIHD_Dictionary(directory)

targetExcel = #commandLineArgs[2]


#pass in target excel file, data dictionary, fills in excel file
def updateExcel(targetExcelFile, parsedData):

    workbook = load_workbook(targetExcelFile)
    sheet = workbook.active

    maxRow = sheet.max_row

    for key in parsedData:
        matchFound = False
        
        for i in range(maxRow):
            actualCell = i + 1
            actualPO = 'A' + str(actualCell)
            actualIHD = 'B' + str(actualCell)

            if int(key) == sheet[actualPO].value:
                sheet[actualIHD] = parsedData[key]
                matchFound = True
                break

        if matchFound == False:
            nextEmptyCellNum = maxRow + 1
            emptyPO = 'A' + str(nextEmptyCellNum)
            emptyIHD = 'B' + str(nextEmptyCellNum)
            sheet[emptyPO] = int(key)
            sheet[emptyIHD] = parsedData[key]
            maxRow += 1
    
    workbook.save(filename= targetExcelFile)

#call update function
updateExcel(targetExcel, parsedData)

